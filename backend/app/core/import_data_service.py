"""Import Data Service

Backend parsing and merge logic for import_data dialog.
"""
from __future__ import annotations

import csv
import json
import os
import uuid
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl

from app.core.pdvm_datenbank import PdvmDatabase


_ALLOWED_TABLES = {"sys_ext_table", "sys_ext_table_man"}


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    if isinstance(value, (list, dict)) and len(value) == 0:
        return True
    return False


def _normalize_header(value: Any) -> str:
    s = str(value or "").strip()
    if not s:
        return ""
    return s.replace(" ", "").replace("-", "").replace("_", "").upper()


def _column_entries(columns: Dict[str, Any]) -> List[Tuple[str, Dict[str, Any]]]:
    entries: List[Tuple[str, Dict[str, Any]]] = []
    for key, cfg in columns.items():
        if isinstance(cfg, dict) and ("key" in cfg or "label" in cfg):
            canon = str(cfg.get("key") or cfg.get("label") or key).strip()
            entries.append((canon, cfg))
        else:
            entries.append((str(key), cfg if isinstance(cfg, dict) else {}))
    return entries


def _build_alias_map(columns: Dict[str, Any], overrides: Optional[Dict[str, str]] = None) -> Dict[str, List[str]]:
    mapping: Dict[str, List[str]] = {}
    for canon, cfg in _column_entries(columns):
        canon_key = _normalize_header(canon)
        aliases: List[str] = [canon_key]
        if overrides:
            override = overrides.get(canon)
            if override:
                override_key = _normalize_header(override)
                if override_key:
                    aliases.insert(0, override_key)
        if isinstance(cfg, dict):
            for alias in cfg.get("aliases") or []:
                alias_key = _normalize_header(alias)
                if alias_key:
                    aliases.append(alias_key)
        mapping[canon] = list(dict.fromkeys(aliases))
    return mapping


def _normalize_match_keys(match_keys: Iterable[str], columns: Dict[str, Any]) -> List[str]:
    alias_map = _build_alias_map(columns)
    alias_lookup: Dict[str, str] = {}
    for canon, aliases in alias_map.items():
        for alias in aliases:
            alias_lookup[alias] = canon

    guid_lookup: Dict[str, str] = {}
    for guid, cfg in columns.items():
        if isinstance(cfg, dict) and ("key" in cfg or "label" in cfg):
            canon = str(cfg.get("key") or cfg.get("label") or guid).strip()
            guid_lookup[str(guid)] = canon

    normalized: List[str] = []
    for key in match_keys or []:
        raw = str(key).strip()
        if not raw:
            continue
        if raw in guid_lookup:
            canon = guid_lookup[raw]
        else:
            canon = alias_lookup.get(_normalize_header(raw), raw)
        if canon and canon not in normalized:
            normalized.append(canon)
    return normalized


def _map_rows(
    headers: List[str],
    rows: List[List[Any]],
    columns_cfg: Dict[str, Any],
    header_overrides: Optional[Dict[str, str]] = None,
) -> Tuple[List[str], List[Dict[str, Any]], List[str]]:
    header_norm = [_normalize_header(h) for h in headers]
    header_index = {h: i for i, h in enumerate(header_norm) if h}

    alias_map = _build_alias_map(columns_cfg, overrides=header_overrides)
    canonical_headers = [canon for canon, _ in _column_entries(columns_cfg)]
    mapped_rows: List[Dict[str, Any]] = []

    for row in rows:
        data: Dict[str, Any] = {}
        for canon, aliases in alias_map.items():
            value = None
            for alias in aliases:
                idx = header_index.get(alias)
                if idx is not None and idx < len(row):
                    value = row[idx]
                    break
            data[canon] = value
        mapped_rows.append(data)

    unmatched = [h for h in header_norm if h and h not in {a for aliases in alias_map.values() for a in aliases}]
    return canonical_headers, mapped_rows, unmatched


def _read_csv_rows(
    path: str,
    limit: int,
    has_headers: bool,
    custom_headers: Optional[List[str]],
) -> Tuple[List[str], List[List[Any]]]:
    with open(path, "rb") as f:
        raw = f.read()
    try:
        text = raw.decode("utf-8-sig")
    except Exception:
        text = raw.decode("latin-1")

    sample = text[:4096]
    dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t"])
    reader = csv.reader(text.splitlines(), dialect)
    rows = list(reader)
    if not rows:
        return [], []
    if has_headers:
        headers = rows[0]
        data_rows = rows[1:limit + 1]
    else:
        headers = custom_headers or []
        if not headers:
            max_len = max((len(r) for r in rows), default=0)
            headers = [f"COL_{i + 1}" for i in range(max_len)]
        data_rows = rows[:limit]
    return headers, data_rows


def _read_xlsx_rows(
    path: str,
    limit: int,
    sheet_name: Optional[str] = None,
    has_headers: bool = True,
    custom_headers: Optional[List[str]] = None,
) -> Tuple[List[str], List[List[Any]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        headers: List[str] = []
        data_rows: List[List[Any]] = []
        for i, row in enumerate(rows_iter):
            if i == 0 and has_headers:
                headers = ["" if v is None else str(v) for v in row]
                continue
            if has_headers and i > limit:
                break
            if not has_headers and i >= limit:
                break
            data_rows.append(list(row))

        if not has_headers:
            headers = custom_headers or []
            if not headers:
                max_len = max((len(r) for r in data_rows), default=0)
                headers = [f"COL_{i + 1}" for i in range(max_len)]
        return headers, data_rows
    finally:
        wb.close()


def _detect_format(filename: str) -> str:
    ext = os.path.splitext(filename or "")[1].lower()
    if ext in {".xlsx", ".xlsm"}:
        return "xlsx"
    if ext in {".csv"}:
        return "csv"
    return "unknown"


def _ensure_allowed_table(table_name: str) -> None:
    if table_name not in _ALLOWED_TABLES:
        raise ValueError("Ungültige Zieltabelle (nur sys_ext_table oder sys_ext_table_man erlaubt)")


def _merge_rows(
    existing: Dict[str, Dict[str, Any]],
    incoming: List[Dict[str, Any]],
    match_keys: List[str],
    conflict_policy: str,
    conflict_rules: Dict[str, str],
    conflict_marker_field: Optional[str],
) -> Dict[str, Dict[str, Any]]:
    index: Dict[Tuple[str, str], str] = {}
    for uid, row in existing.items():
        for key in match_keys:
            val = row.get(key)
            if _is_empty(val):
                continue
            index_key = (key, str(val).strip().lower())
            index.setdefault(index_key, uid)

    def apply_merge(base: Dict[str, Any], update: Dict[str, Any]) -> Dict[str, Any]:
        result = dict(base)
        if conflict_policy == "update_wins":
            for k, v in update.items():
                if not _is_empty(v):
                    result[k] = v
            return result
        if conflict_policy == "base_wins":
            for k, v in update.items():
                if _is_empty(result.get(k)) and not _is_empty(v):
                    result[k] = v
            return result
        if conflict_policy == "field_priority":
            for k, v in update.items():
                rule = str(conflict_rules.get(k) or "").strip().lower()
                if rule == "base":
                    continue
                if rule == "update" and not _is_empty(v):
                    result[k] = v
                if rule == "if_empty" and _is_empty(result.get(k)) and not _is_empty(v):
                    result[k] = v
            return result
        return result

    updated = dict(existing)

    for row in incoming:
        matched_uid = None
        for key in match_keys:
            val = row.get(key)
            if _is_empty(val):
                continue
            matched_uid = index.get((key, str(val).strip().lower()))
            if matched_uid:
                break

        if matched_uid:
            if conflict_policy == "insert_new_only":
                continue
            if conflict_policy == "new_record_on_conflict":
                new_uid = str(uuid.uuid4())
                new_row = dict(row)
                if conflict_marker_field:
                    new_row[conflict_marker_field] = f"conflict_with:{matched_uid}"
                updated[new_uid] = new_row
                continue

            base = updated.get(matched_uid, {})
            updated[matched_uid] = apply_merge(base, row)
        else:
            new_uid = str(uuid.uuid4())
            updated[new_uid] = dict(row)

    return updated


async def load_dataset(gcs, *, table_name: str, dataset_uid: str) -> Dict[str, Any]:
    _ensure_allowed_table(table_name)
    db = PdvmDatabase(table_name, system_pool=gcs._system_pool, mandant_pool=gcs._mandant_pool)
    row = await db.get_by_uid(uuid.UUID(str(dataset_uid)))
    if not row:
        raise KeyError("Datensatz nicht gefunden")
    daten = row.get("daten") or {}
    return {
        "uid": str(row.get("uid")),
        "name": row.get("name") or "",
        "daten": daten,
    }


def parse_file_to_preview(
    *,
    path: str,
    filename: str,
    columns_cfg: Dict[str, Any],
    sheet_name: Optional[str] = None,
    has_headers: bool = True,
    custom_headers: Optional[List[str]] = None,
    header_overrides: Optional[Dict[str, str]] = None,
    limit: int = 200,
) -> Dict[str, Any]:
    file_format = _detect_format(filename)
    if file_format == "csv":
        headers, rows = _read_csv_rows(path, limit, has_headers=has_headers, custom_headers=custom_headers)
    elif file_format == "xlsx":
        headers, rows = _read_xlsx_rows(
            path,
            limit,
            sheet_name=sheet_name,
            has_headers=has_headers,
            custom_headers=custom_headers,
        )
    else:
        raise ValueError("Dateiformat nicht unterstuetzt (nur CSV/XLSX)")

    if not headers:
        raise ValueError("Keine Header gefunden")

    canonical_headers, mapped_rows, unmatched = _map_rows(
        headers,
        rows,
        columns_cfg,
        header_overrides=header_overrides,
    )

    return {
        "file_format": file_format,
        "headers": headers,
        "canonical_headers": canonical_headers,
        "rows": mapped_rows,
        "unmatched_headers": unmatched,
    }


async def apply_preview_rows(
    gcs,
    *,
    table_name: str,
    dataset_uid: str,
    rows: List[Dict[str, Any]],
) -> Dict[str, Any]:
    dataset = await load_dataset(gcs, table_name=table_name, dataset_uid=dataset_uid)
    daten = dataset.get("daten") if isinstance(dataset.get("daten"), dict) else {}
    root = daten.get("ROOT") if isinstance(daten.get("ROOT"), dict) else {}
    config = daten.get("CONFIG") if isinstance(daten.get("CONFIG"), dict) else {}
    columns_cfg = config.get("COLUMNS") if isinstance(config.get("COLUMNS"), dict) else {}

    match_keys = [str(k) for k in root.get("MATCH_KEYS") or []]
    if not match_keys:
        match_keys = [str(k) for k in config.get("KEY_MERGE_PRIORITY") or []]
    match_keys = _normalize_match_keys(match_keys, columns_cfg)

    conflict_policy = str(root.get("CONFLICT_POLICY") or "base_wins").strip().lower()
    conflict_rules = root.get("CONFLICT_RULES") if isinstance(root.get("CONFLICT_RULES"), dict) else {}
    conflict_marker_field = root.get("CONFLICT_MARKER_FIELD")

    existing = daten.get("DATAS") if isinstance(daten.get("DATAS"), dict) else {}

    if any(isinstance(r, dict) and "__uid" in r for r in rows):
        merged = {}
        for row in rows:
            if not isinstance(row, dict):
                continue
            uid = str(row.get("__uid") or uuid.uuid4())
            merged[uid] = {k: v for k, v in row.items() if k != "__uid"}
    else:
        merged = _merge_rows(
            existing=existing,
            incoming=rows,
            match_keys=match_keys,
            conflict_policy=conflict_policy,
            conflict_rules=conflict_rules,
            conflict_marker_field=conflict_marker_field,
        )

    daten = dict(daten)
    daten["DATAS"] = merged
    root = dict(root)
    root["LAST_IMPORT_AT"] = datetime.utcnow().isoformat() + "Z"
    daten["ROOT"] = root

    db = PdvmDatabase(table_name, system_pool=gcs._system_pool, mandant_pool=gcs._mandant_pool)
    await db.update(uuid.UUID(str(dataset_uid)), daten=daten, name=dataset.get("name"))

    return {
        "updated_count": len(merged),
        "dataset_uid": dataset_uid,
    }


